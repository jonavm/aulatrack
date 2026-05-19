from __future__ import annotations

import csv
import html
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QMarginsF, QSizeF
from PySide6.QtGui import QPageLayout, QPageSize, QPagedPaintDevice, QPdfWriter, QTextDocument


class ExportService:
    def export_table(
        self,
        destination: str,
        headers: list[str],
        rows: list[list[str]],
        metadata: list[tuple[str, str]] | None = None,
        title: str = "Exportacion AulaTrack",
    ) -> Path:
        target = Path(destination).expanduser()
        suffix = target.suffix.lower()
        if suffix == ".pdf":
            self._export_pdf(target, headers, rows, metadata or [], title)
            return target
        if suffix == ".xlsx":
            self._export_xlsx(target, headers, rows, metadata or [], title)
            return target

        if suffix != ".csv":
            target = target.with_suffix(".csv")
        self._export_csv(target, headers, rows, metadata or [], title)
        return target

    @staticmethod
    def _export_csv(
        target: Path,
        headers: list[str],
        rows: list[list[str]],
        metadata: list[tuple[str, str]],
        title: str,
    ) -> None:
        target.parent.mkdir(parents=True, exist_ok=True)
        with open(target, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow([title])
            writer.writerow([])
            for label, value in metadata:
                writer.writerow([label, value])
            if metadata:
                writer.writerow([])
            writer.writerow(headers)
            writer.writerows(rows)

    @staticmethod
    def _export_xlsx(
        target: Path,
        headers: list[str],
        rows: list[list[str]],
        metadata: list[tuple[str, str]],
        title: str,
    ) -> None:
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Exportacion"

        header_fill = PatternFill("solid", fgColor="E7EEF7")
        title_fill = PatternFill("solid", fgColor="DCE9FA")
        thin_border = Border(
            left=Side(style="thin", color="D6DFEA"),
            right=Side(style="thin", color="D6DFEA"),
            top=Side(style="thin", color="D6DFEA"),
            bottom=Side(style="thin", color="D6DFEA"),
        )

        current_row = 1
        sheet.cell(row=current_row, column=1, value=title)
        sheet.cell(row=current_row, column=1).font = Font(size=14, bold=True)
        sheet.cell(row=current_row, column=1).fill = title_fill
        current_row += 2

        for label, value in metadata:
            sheet.cell(row=current_row, column=1, value=label)
            sheet.cell(row=current_row, column=2, value=value)
            sheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

        if metadata:
            current_row += 1

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        current_row += 1

        for row_values in rows:
            for col, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=current_row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
            current_row += 1

        for col_index, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for row_values in rows:
                if col_index - 1 < len(row_values):
                    max_length = max(max_length, len(str(row_values[col_index - 1])))
            sheet.column_dimensions[get_column_letter(col_index)].width = min(max_length + 3, 28)

        workbook.save(target)

    @staticmethod
    def _export_pdf(
        target: Path,
        headers: list[str],
        rows: list[list[str]],
        metadata: list[tuple[str, str]],
        title: str,
    ) -> None:
        target.parent.mkdir(parents=True, exist_ok=True)
        writer = QPdfWriter(str(target))
        writer.setResolution(96)
        writer.setPageOrientation(QPageLayout.Landscape)
        writer.setPageSize(QPageSize(QPageSize.A4))
        writer.setPageMargins(QMarginsF(12, 12, 12, 12), QPageLayout.Millimeter)
        writer.setTitle("Exportacion AulaTrack")

        page_rect = writer.pageLayout().paintRect(QPageLayout.Point)
        document = QTextDocument()
        document.setPageSize(QSizeF(page_rect.width(), page_rect.height()))
        document.setHtml(ExportService._build_html(headers, rows, metadata, title))
        document.print_(writer)

        if not target.exists() or target.stat().st_size == 0:
            raise OSError("No se pudo crear el archivo PDF.")

    @staticmethod
    def _build_html(
        headers: list[str],
        rows: list[list[str]],
        metadata: list[tuple[str, str]],
        title: str,
    ) -> str:
        metadata_html = ""
        if metadata:
            metadata_rows = "".join(
                f"<tr><td class='meta-label'>{html.escape(label)}</td><td>{html.escape(value)}</td></tr>"
                for label, value in metadata
            )
            metadata_html = f"<table class='meta'>{metadata_rows}</table>"

        header_html = "".join(f"<th>{html.escape(header)}</th>" for header in headers)
        body_html = "".join(
            "<tr>" + "".join(f"<td>{html.escape(str(value))}</td>" for value in row) + "</tr>"
            for row in rows
        )

        return f"""
        <html>
            <head>
                <meta charset="utf-8">
                <style>
                    body {{
                        font-family: 'Segoe UI', sans-serif;
                        font-size: 10pt;
                        color: #142033;
                    }}
                    h1 {{
                        font-size: 18pt;
                        margin: 0 0 6px 0;
                        color: #1C2E49;
                    }}
                    .subtitle {{
                        font-size: 9pt;
                        color: #6E7C91;
                        margin-bottom: 14px;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin-bottom: 14px;
                    }}
                    .meta {{
                        width: auto;
                    }}
                    .meta td {{
                        border: none;
                        padding: 2px 8px 2px 0;
                    }}
                    .meta-label {{
                        font-weight: 700;
                        color: #274C77;
                    }}
                    th {{
                        background: #E7EEF7;
                        color: #142033;
                        font-weight: 700;
                        border: 1px solid #C8D6E8;
                        padding: 6px;
                    }}
                    td {{
                        border: 1px solid #D6DFEA;
                        padding: 5px 6px;
                    }}
                    tr:nth-child(even) td {{
                        background: #F7FAFC;
                    }}
                </style>
            </head>
            <body>
                <h1>{html.escape(title)}</h1>
                <div class="subtitle">Generado desde AulaTrack</div>
                {metadata_html}
                <table>
                    <thead>
                        <tr>{header_html}</tr>
                    </thead>
                    <tbody>
                        {body_html}
                    </tbody>
                </table>
            </body>
        </html>
        """
