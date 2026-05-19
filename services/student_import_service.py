from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from models.entities import Group


@dataclass(slots=True)
class ParsedStudent:
    first_name: str
    last_name: str
    student_code: str = ""


@dataclass(slots=True)
class ImportPreview:
    source_name: str
    total_rows: int
    detected_mode: str
    students: list[ParsedStudent]
    warnings: list[str]
    detected_group: str = ""
    detected_school: str = ""
    detected_school_year: str = ""
    sheet_name: str = ""


class StudentImportService:
    NAME_KEYS = {"nombre", "nombres", "name", "first_name", "nombre(s)"}
    LAST_NAME_KEYS = {"apellido", "apellidos", "last_name", "surname", "apellidos(s)"}
    FULL_NAME_KEYS = {
        "alumno",
        "nombre completo",
        "full_name",
        "student",
        "estudiante",
        "nombre del alumno (a)",
        "nombre del alumno(a)",
        "nombre del alumno",
    }
    CODE_KEYS = {"matricula", "codigo", "clave", "student_code", "no_lista", "numero"}

    def parse_file(self, file_path: str) -> ImportPreview:
        path = Path(file_path)
        suffix = path.suffix.lower()
        sheet_name = ""
        if suffix == ".csv":
            rows = self._read_csv(path)
        elif suffix in {".xlsx", ".xlsm"}:
            rows, sheet_name = self._read_xlsx(path)
        else:
            raise ValueError("Formato no soportado. Usa CSV o XLSX.")

        if not rows:
            raise ValueError("El archivo no contiene filas utilizables.")

        header_row_index = self._detect_header_row(rows)
        headers = [self._normalize_header(value) for value in rows[header_row_index]]
        data_rows = rows[header_row_index + 1 :]

        first_name_index = self._find_header_index(headers, self.NAME_KEYS)
        last_name_index = self._find_header_index(headers, self.LAST_NAME_KEYS)
        full_name_index = self._find_header_index(headers, self.FULL_NAME_KEYS)
        code_index = self._find_header_index(headers, self.CODE_KEYS)

        warnings: list[str] = []
        students: list[ParsedStudent] = []
        detected_school, detected_school_year, detected_group = self._extract_metadata(rows)

        if first_name_index is not None and last_name_index is not None:
            detected_mode = "columnas separadas"
            for row in data_rows:
                first_name = self._cell(row, first_name_index)
                last_name = self._cell(row, last_name_index)
                if not first_name and not last_name:
                    continue
                if not first_name or not last_name:
                    warnings.append("Algunas filas no tienen nombre o apellidos completos.")
                    continue
                students.append(
                    ParsedStudent(
                        first_name=first_name,
                        last_name=last_name,
                        student_code=self._cell(row, code_index) if code_index is not None else "",
                    )
                )
        elif full_name_index is not None:
            detected_mode = "nombre completo"
            for row in data_rows:
                full_name = self._cell(row, full_name_index)
                if not full_name:
                    continue
                first_name, last_name = self._split_full_name(full_name)
                if not first_name or not last_name:
                    warnings.append("Algunas filas no pudieron separarse claramente en nombre y apellidos.")
                    continue
                students.append(
                    ParsedStudent(
                        first_name=first_name,
                        last_name=last_name,
                        student_code=self._cell(row, code_index) if code_index is not None else "",
                    )
                )
        else:
            inferred = self._infer_student_column(data_rows)
            if inferred is None:
                raise ValueError(
                    "No pude detectar columnas de alumno. Usa encabezados como Nombre, Apellidos o Alumno."
                )
            detected_mode = "deteccion automatica por contenido"
            inferred_name_index, inferred_code_index = inferred
            for row in data_rows:
                full_name = self._cell(row, inferred_name_index)
                if not self._looks_like_student_name(full_name):
                    continue
                first_name, last_name = self._split_full_name(full_name)
                if not first_name or not last_name:
                    warnings.append("Algunas filas no pudieron separarse claramente en nombre y apellidos.")
                    continue
                students.append(
                    ParsedStudent(
                        first_name=first_name,
                        last_name=last_name,
                        student_code=self._cell(row, inferred_code_index) if inferred_code_index is not None else "",
                    )
                )

        if not students:
            raise ValueError("No se detectaron alumnos validos para importar.")

        return ImportPreview(
            source_name=path.name,
            total_rows=len(data_rows),
            detected_mode=detected_mode,
            students=students,
            warnings=warnings,
            detected_group=detected_group,
            detected_school=detected_school,
            detected_school_year=detected_school_year,
            sheet_name=sheet_name,
        )

    @staticmethod
    def split_group_label(group_label: str) -> tuple[str, str]:
        normalized = StudentImportService._normalize_group_label(group_label)
        if not normalized:
            return "", ""

        separators = ("-",)
        for separator in separators:
            if separator in normalized:
                left, right = normalized.split(separator, 1)
                return Group.normalize_grade_level(left.strip()), right.strip().upper()

        if len(normalized) >= 2:
            return Group.normalize_grade_level(normalized[:-1].strip()), normalized[-1].strip().upper()
        return Group.normalize_grade_level(normalized.strip()), ""

    def _detect_header_row(self, rows: list[list[str]]) -> int:
        best_index = -1
        best_score = -1

        for index, row in enumerate(rows[:15]):
            normalized = [self._normalize_header(value) for value in row if value.strip()]
            if not normalized:
                continue

            score = 0
            for value in normalized:
                if value in self.NAME_KEYS:
                    score += 3
                if value in self.LAST_NAME_KEYS:
                    score += 3
                if value in self.FULL_NAME_KEYS:
                    score += 4
                if value in self.CODE_KEYS:
                    score += 1

            if score > best_score:
                best_score = score
                best_index = index

        if best_index < 0 or best_score <= 0:
            raise ValueError(
                "No pude detectar columnas de alumno. Usa encabezados como Nombre, Apellidos o Alumno."
            )

        return best_index

    def _infer_student_column(self, rows: list[list[str]]) -> tuple[int, int | None] | None:
        sample_rows = rows[:80]
        max_cols = max((len(row) for row in sample_rows), default=0)
        best_index = None
        best_score = 0

        for column_index in range(max_cols):
            score = 0
            for row in sample_rows:
                value = self._cell(row, column_index)
                if self._looks_like_student_name(value):
                    score += 1
            if score > best_score:
                best_score = score
                best_index = column_index

        if best_index is None or best_score < 3:
            return None

        code_index = None
        for candidate in range(max_cols):
            if candidate == best_index:
                continue
            numeric_like = 0
            for row in sample_rows[:40]:
                value = self._cell(row, candidate)
                if value and value.replace(".", "", 1).isdigit():
                    numeric_like += 1
            if numeric_like >= 3:
                code_index = candidate
                break

        return best_index, code_index

    def _read_csv(self, path: Path) -> list[list[str]]:
        encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
        for encoding in encodings:
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    return [list(row) for row in csv.reader(handle) if any(cell.strip() for cell in row)]
            except UnicodeDecodeError:
                continue
        raise ValueError("No pude leer el CSV por codificacion.")

    def _read_xlsx(self, path: Path) -> tuple[list[list[str]], str]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(values):
                rows.append(values)
        return rows, sheet.title

    @staticmethod
    def _normalize_header(value: str) -> str:
        return (
            value.strip()
            .lower()
            .replace(".", "")
            .replace("_", " ")
            .replace("-", " ")
        )

    @staticmethod
    def _find_header_index(headers: list[str], candidates: set[str]) -> int | None:
        for index, header in enumerate(headers):
            if header in candidates:
                return index
        return None

    @staticmethod
    def _cell(row: list[str], index: int | None) -> str:
        if index is None or index >= len(row):
            return ""
        return str(row[index]).strip()

    @staticmethod
    def _split_full_name(full_name: str) -> tuple[str, str]:
        cleaned = StudentImportService._clean_full_name(full_name)
        if "/" in cleaned and "*" in cleaned:
            return StudentImportService._split_public_school_name(cleaned)
        if "*" in cleaned:
            left, right = [part.strip() for part in cleaned.split("*", 1)]
            if left and right:
                return " ".join(right.split()), " ".join(piece.strip() for piece in left.split("/") if piece.strip())
            cleaned = left or right

        parts = [part for part in cleaned.split() if part]
        if len(parts) < 2:
            return cleaned.strip(), ""
        if len(parts) == 2:
            return parts[0], parts[1]
        if len(parts) == 3:
            return " ".join(parts[:2]), parts[2]
        return " ".join(parts[:-2]), " ".join(parts[-2:])

    @staticmethod
    def _clean_full_name(full_name: str) -> str:
        cleaned = " ".join(full_name.strip().split())
        cleaned = cleaned.strip(" -*,;")
        return cleaned

    @staticmethod
    def _split_public_school_name(full_name: str) -> tuple[str, str]:
        normalized = " ".join(full_name.strip().split())
        try:
            last_names_part, first_names_part = normalized.split("*", 1)
        except ValueError:
            return normalized, ""

        last_names = " ".join(
            piece.strip()
            for piece in last_names_part.split("/")
            if piece.strip()
        )
        first_names = " ".join(first_names_part.strip().split())
        return first_names, last_names

    @staticmethod
    def _looks_like_student_name(value: str) -> bool:
        text = " ".join(value.strip().split())
        if not text:
            return False
        if len(text) < 5:
            return False
        upper = text.upper()
        if "/" in text and "*" in text:
            return True
        parts = [part for part in text.split() if part]
        if len(parts) >= 2 and not any(char.isdigit() for char in text):
            return True
        if "ALUMNO" in upper or "ESCUELA" in upper or "CURSO" in upper:
            return False
        return False

    @staticmethod
    def _extract_metadata(rows: list[list[str]]) -> tuple[str, str, str]:
        school_name = ""
        school_year = ""
        group_name = ""

        for row in rows[:8]:
            if not row:
                continue
            first = row[0].strip()
            upper_first = first.upper()
            if upper_first.startswith("ESCUELA ") and not school_name:
                school_name = first
            if "CURSO ESCOLAR" in upper_first and not school_year:
                school_year = first.replace("CURSO ESCOLAR", "").strip()

            merged = " ".join(cell.strip() for cell in row if cell.strip())
            upper_merged = merged.upper()
            marker = "GRADO Y GRUPO:"
            if marker in upper_merged and not group_name:
                original = merged
                idx = upper_merged.index(marker) + len(marker)
                raw_group = original[idx:].strip().split()[0]
                group_name = StudentImportService._normalize_group_label(raw_group)

        return school_name, school_year, group_name

    @staticmethod
    def _normalize_group_label(raw_group: str) -> str:
        label = raw_group.strip().replace(".", "-").replace(" ", "")
        parts = label.split("-", 1)
        if len(parts) == 2:
            grade_level = Group.normalize_grade_level(parts[0].strip())
            group_section = parts[1].strip().upper()
            if grade_level and group_section:
                return f"{grade_level}-{group_section}"
        if len(label) >= 2:
            grade_level = Group.normalize_grade_level(label[:-1].strip())
            group_section = label[-1].strip().upper()
            if grade_level and group_section:
                return f"{grade_level}-{group_section}"
        return label
